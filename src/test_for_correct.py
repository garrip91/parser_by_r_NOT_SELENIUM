from datetime import datetime

import openpyxl


book = openpyxl.open('ПРИОРИТЕТ № 1.xlsx')
sheets = book.sheetnames
sheet = book[sheets[0]]

# for cell in sheet['2']:
    # print(cell.value)

# c = 0
# for cell in needed:
    # v = cell.value
    # c += 1
# print(c)

rows = sheet.max_row
cols = sheet.max_column
correct_lines = []
for i in range(2, rows + 1):
    correct_line = []
    for j in range(1, cols + 1):
        cell = sheet.cell(row = i, column = j)
        correct_line.append(str(cell.value).strip())
    correct_lines.append(correct_line)

with open("unique_data_NEW.txt", 'a', encoding='UTF-8') as f2:
    for line in correct_lines:
        f2.write(F'{str(line).strip("[]")}\n')
    


    
    
    
    

    
    
book.close()




# data_set = set()

# with open("unique_data.txt", "r", encoding="UTF-8") as f:
    
    # for line in f.readlines():
        # if line == "\n":
            # continue
        # else:
            # line_list = line.split(' || ')
            # data_set.add(line)
      
      
# data_list = list(data_set)
    
    
# with open("unique_data_selected.txt", "a", encoding="UTF-8") as f:
        
    # for line1 in data_list:
        # if line2 in line1:
            # f.write(line1)
            # del articles_list[articles_list.index(line2)]
            # print(len(articles_list))
        # else:
            # print(len(articles_list))