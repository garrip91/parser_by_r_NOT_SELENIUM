import openpyxl
from openpyxl.styles import Font, Alignment, NamedStyle, Border, Side


book = openpyxl.Workbook()
sheets = book.sheetnames
sheet = book[sheets[0]]

# ТОЛСТЫЕ ГРАНИЦЫ:
thick_border = Border(
    left=Side(style='thick'), 
    right=Side(style='thick'), 
    top=Side(style='thick'), 
    bottom=Side(style='thick')
)

# ТОНКИЕ ГРАНИЦЫ:
thin_border = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    #top=Side(style='thin'), 
    bottom=Side(style='thin')
)

sheet['A1'] = "№ П/П"
sheet['A1'].font = Font(bold=True)
sheet['A1'].alignment = Alignment(horizontal='center')
sheet.column_dimensions['A'].width = 10
sheet['B1'] = "ID ТОВАРА"
sheet['B1'].font = Font(bold=True)
sheet['B1'].alignment = Alignment(horizontal='center')
sheet.column_dimensions['B'].width = 20
sheet['C1'] = "НАЗВАНИЕ ТОВАРА"
sheet['C1'].font = Font(bold=True)
sheet['C1'].alignment = Alignment(horizontal='center')
sheet.column_dimensions['C'].width = 20
######
sheet['D1'] = "КОД ТОВАРА"
sheet['D1'].font = Font(bold=True)
sheet['D1'].alignment = Alignment(horizontal='center')
sheet.column_dimensions['D'].width = 20
######
sheet['E1'] = "ЦЕНА ТОВАРА (руб.)"
sheet['E1'].font = Font(bold=True)
sheet['E1'].alignment = Alignment(horizontal='center')
sheet.column_dimensions['E'].width = 20
sheet['F1'] = "КОЛИЧЕСТВО ТОВАРА (шт./упак.)"
sheet['F1'].font = Font(bold=True)
sheet['F1'].alignment = Alignment(horizontal='center')
sheet.column_dimensions['F'].width = 35

sheet.freeze_panes = 'A2'

area = [
    sheet['A1'],
    sheet['B1'],
    sheet['C1'],
    sheet['D1'],
    sheet['E1'],
    sheet['F1']
]

################# 1. ГРАНИЦЫ ЗАГОЛОВОЧНОЙ СТРОКИ: #################
for cell in area:
    cell.border = thick_border
############################### 1.- ###############################

######### 2. ОТКРЫВАЕМ ИСХОДНЫЙ ФАЙЛ С ДАННЫМИ НА ЧТЕНИЕ: #########
with open('data.txt', 'r', encoding='UTF-8') as source:
############################### 2.- ###############################

    list_for_data_count = []    
    with open("data.txt", "r", encoding="UTF-8") as f:
        for line in f.readlines():
            if line == "\n":
                continue
            else:
                list_for_data_count.append(line.split(' || '))
        list_for_data_count = set(list_for_NEW_data_count)

    list_for_data_articles_count = []
    for i in list_for_data_count:
        list_for_data_articles_count.append(i[2])
    list_for_data_articles_count = set(list_for_data_articles_count)
    
    list_for_NEW_data_count = []
    for i in list_for_data_count:
        if i[2] in list_for_data_articles_count:
            if 

    ######## 3. ЦИКЛОМ ЗАПИСЫВАЕМ ДАННЫЕ ПО ВСЕМ ТОВАРАМ, ВКЛЮЧАЯ ПОРЯДКОВЫЙ НОМЕР КАЖДОГО ТОВАРА: ########
    nums = 2
    letters = ['A', 'B', 'C', 'D', 'E', 'F']
    n_p_p = 1
    nums_insert = 1
    for line in source:
        line = line.replace('\n', '')
        result = line.split(' || ')
        result.insert(0, nums_insert)
        #print(result)
        nums_insert += 1
        letters_index = 0
        for elem in result:
            #print(elem)
        #print("========================================================")
            sheet[F'{letters[letters_index]}{nums}'] = elem
            # sheet[F'{letters[letters_index]}{nums}'].alignment = Alignment(horizontal='center', vertical='center')
            # sheet[F'{letters[letters_index]}{nums}'].alignment = Alignment(wrap_text=True)
            letters_index += 1
        nums += 1
        
    n_p_p_count = sheet.max_row-1
    #print(n_p_p_count)
    ################################################# 3.- #################################################

    ############################ 4. ВЫДЕЛЯЕМ ЗАГОЛОВОЧНУЮ СТРОКУ ЖИРНЫМ ШРИФТОМ: ##########################
    n = 2
    for i in range(n_p_p_count):
        #print(type(i))
        #print(i)
        sheet[F'A{n}'].font = Font(bold=True)
        n += 1
    ################################################# 4.- #################################################
                    
    ################## 5. УДАЛЯЕМ ВСЕ НАШИ ПЕРЕМЕННЫЕ ИЗ ОЗУ: ##################
    del nums
    del letters
    del n_p_p
    del nums_insert
    del letters_index
    ################################ 5.- #######################################        
    
    ######################### 6. ВЫРАВНИВАЕМ ПО ЦЕНТРУ СОДЕРЖИМОЕ КАЖДОЙ ЯЧЕЙКИ: ##########################
    nums = 2
    letters = ['A', 'B', 'C', 'D', 'E', 'F']
    n_p_p = 1
    nums_insert = 1
    #for i in range(sheet.max_row):
    for i in range(sheet.max_column):
        nums_insert += 1
        letters_index = 0
        for j in range(len(letters)):
            #sheet[F'{letters[letters_index]}{nums}'].alignment = Alignment(horizontal='center')
            sheet[F'{letters[letters_index]}{nums}'].alignment = Alignment(wrap_text=True)
            #sheet[F'{letters[letters_index]}{nums}'].alignment = Alignment(vertical='center')
#            sheet[F'A{letters[letters_index]}{nums}'].border = thin_border
            letters_index += 1
        nums += 1
    #######################################################################################################
            
    ############### 5. СНОВА УДАЛЯЕМ ВСЕ НАШИ ПЕРЕМЕННЫЕ ИЗ ОЗУ: ###############
    del nums
    del letters
    del n_p_p
    del nums_insert
    del letters_index
    ################################ 5.- #######################################        


# sheet['A20'].border = thin_border - пример для одной ячейки

book.save('ТОВАРЫ_ВАШЕГО_ПОСТАВЩИКА.xlsx')
book.close()